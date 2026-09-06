"""推荐校准数据准备与数据包契约。

独立于推荐参数推断：负责原始观测解析、固定划分、训练侧筛选/缩放、
分层抽样、数据包校验与原子发布。不导入重量级推断模块或 run_analysis。

数据包由 train.json / test.json / manifest.json 组成：
- train.json / test.json：schema_version=1、split、num_agents、scale_ratio、records；
- manifest.json：schema_version=1、源文件散列、字段映射、清洗计数、划分种子/比例、
  两分区 ID、选择策略、锚点/比例、人口规模、两分区记录数与文件散列。
"""

from __future__ import annotations

import csv
import hashlib
import json
import math
import os
import shutil
import uuid
from contextlib import contextmanager
from pathlib import Path
from typing import Any, Dict, Iterator, List, Optional, Sequence, Tuple

import numpy as np

SCHEMA_VERSION = 1

# 训练过滤阈值（与方案 §4 一致）
MIN_VIEW_FOR_FILTER = 100.0
MIN_REPOST_FOR_FILTER = 0.0  # 严格大于 0

_REQUIRED_RECORD_FIELDS = (
    "story_id",
    "text",
    "repost_count",
    "view_count",
    "scaled_target",
    "unclipped_target",
    "target_clipped",
)


def file_sha256(path: Path) -> str:
    """计算文件 SHA-256（十六进制）。"""
    digest = hashlib.sha256()
    with Path(path).open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _read_json(path: Path) -> Dict[str, Any]:
    """读取 JSON 对象（load_split 与测试共用，便于拦截文件访问）。"""
    with Path(path).open("r", encoding="utf-8") as f:
        payload = json.load(f)
    if not isinstance(payload, dict):
        raise ValueError(f"文件 {path} 必须是 JSON 对象。")
    return payload


def _write_json(path: Path, payload: Dict[str, Any]) -> None:
    """写入 JSON（禁止 NaN/Infinity）。"""
    with Path(path).open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2, allow_nan=False)


# ---------------------------------------------------------------
# 原子发布
# ---------------------------------------------------------------

def _lock_path_for(output_dir: Path) -> Path:
    return output_dir.parent / (output_dir.name + ".lock")


def _remove_lock(lock_path: Path) -> None:
    try:
        lock_path.unlink()
    except FileNotFoundError:
        pass


@contextmanager
def publish_output(output_dir: Path) -> Iterator[Path]:
    """原子发布上下文管理器。

    向调用者提供同父目录下的临时目录；成功退出后整体重命名为 output_dir。
    使用独占创建的同名锁文件防止并发：已存在输出或锁即失败。硬杀遗留锁
    报错并提示核实进程后人工移除，不自动抢锁。
    """
    output_dir = Path(output_dir)
    parent = output_dir.parent
    parent.mkdir(parents=True, exist_ok=True)

    lock_path = _lock_path_for(output_dir)
    try:
        lock_fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
    except FileExistsError as exc:
        pid_hint = "未知"
        try:
            pid_hint = lock_path.read_text(encoding="utf-8").strip()
        except OSError:
            pass
        raise FileExistsError(
            f"输出目录被占用（锁已存在，进程 {pid_hint}）: {lock_path}。"
            f"请核实该进程后人工移除锁文件再重试。"
        ) from exc
    with os.fdopen(lock_fd, "w", encoding="utf-8") as lf:
        lf.write(str(os.getpid()))

    if output_dir.exists():
        _remove_lock(lock_path)
        raise FileExistsError(f"输出目录已存在: {output_dir}")

    temp_dir = parent / f".{output_dir.name}.tmp-{os.getpid()}-{uuid.uuid4().hex[:8]}"
    temp_dir.mkdir()
    try:
        yield temp_dir
        if output_dir.exists():
            raise FileExistsError(f"输出目录已存在: {output_dir}")
        os.replace(temp_dir, output_dir)
    except BaseException:
        shutil.rmtree(temp_dir, ignore_errors=True)
        raise
    finally:
        _remove_lock(lock_path)


# ---------------------------------------------------------------
# 值规范化与校验
# ---------------------------------------------------------------

def _normalize_id(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "1" if value else "0"
    if isinstance(value, float):
        if math.isnan(value):
            return ""
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value)


def _parse_number(value: Any, *, field: str, row_no: int) -> float:
    """解析非负有界数值；空值视为 0，无法解析/非有限/负数报错。"""
    if value is None:
        return 0.0
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        number = float(value)
    elif isinstance(value, str):
        text = value.strip()
        if text == "":
            return 0.0
        try:
            number = float(text)
        except ValueError as exc:
            raise ValueError(
                f"第 {row_no} 行字段 {field!r} 无法解析为数值: {value!r}"
            ) from exc
    else:
        raise ValueError(
            f"第 {row_no} 行字段 {field!r} 类型不支持: {type(value).__name__}"
        )
    if math.isnan(number) or math.isinf(number):
        raise ValueError(f"第 {row_no} 行字段 {field!r} 必须是有界数值，得到 {value!r}")
    if number < 0:
        raise ValueError(f"第 {row_no} 行字段 {field!r} 必须非负，得到 {value!r}")
    return number


# ---------------------------------------------------------------
# 源读取
# ---------------------------------------------------------------

def _read_json_records(source: Path) -> List[Dict[str, Any]]:
    payload = _read_json(source)
    records = payload.get("records")
    if not isinstance(records, list):
        raise ValueError(f"JSON 输入 {source} 需要提供 records 数组。")
    result: List[Dict[str, Any]] = []
    for index, item in enumerate(records):
        row_no = index + 1
        if not isinstance(item, dict):
            raise ValueError(f"records[{index}] 必须是 JSON 对象。")
        for field in ("story_id", "text", "repost_count", "view_count"):
            if field not in item:
                raise ValueError(f"records[{index}] 缺少字段：{field}")
        story_id = _normalize_id(item.get("story_id"))
        text = _normalize_text(item.get("text"))
        repost = _parse_number(item.get("repost_count"), field="repost_count", row_no=row_no)
        views = _parse_number(item.get("view_count"), field="view_count", row_no=row_no)
        result.append(
            {
                "story_id": story_id,
                "text": text,
                "repost_count": repost,
                "view_count": views,
            }
        )
    return result


def _read_table_rows(source: Path) -> List[Dict[str, Any]]:
    """读取 CSV/XLSX 为 list[dict]，值为原始单元格类型。"""
    suffix = source.suffix.lower()
    if suffix == ".csv":
        with source.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            return [dict(row) for row in reader]
    if suffix == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ModuleNotFoundError as exc:
            raise ModuleNotFoundError("缺少 openpyxl，请先执行 uv sync。") from exc
        workbook = load_workbook(source, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()
        if not rows:
            return []
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        result: List[Dict[str, Any]] = []
        for row in rows[1:]:
            item = {
                header[index]: row[index] if index < len(row) else None
                for index in range(len(header))
                if header[index]
            }
            result.append(item)
        return result
    raise ValueError(f"暂不支持的表格格式：{source.suffix}")


def _read_table_records(
    source: Path,
    *,
    text_column: Optional[str],
    retweet_columns: str,
    view_column: str,
    id_column: str,
) -> List[Dict[str, Any]]:
    if not text_column:
        raise ValueError("表格输入必须提供 --text-column。")
    rows = _read_table_rows(source)
    retweet_fields = [item.strip() for item in retweet_columns.split(",") if item.strip()]
    if not retweet_fields:
        raise ValueError("retweet_columns 不能为空。")

    if not rows:
        raise ValueError("表格为空，无法准备数据包。")
    available = set(rows[0].keys())
    missing = [c for c in [id_column, text_column, view_column, *retweet_fields] if c not in available]
    if missing:
        raise ValueError(f"表格缺列：{missing}")

    result: List[Dict[str, Any]] = []
    for index, row in enumerate(rows):
        row_no = index + 2  # 表头为第 1 行
        story_id = _normalize_id(row.get(id_column))
        text = _normalize_text(row.get(text_column))
        repost = sum(
            _parse_number(row.get(field), field=field, row_no=row_no)
            for field in retweet_fields
        )
        views = _parse_number(row.get(view_column), field=view_column, row_no=row_no)
        result.append(
            {
                "story_id": story_id,
                "text": text,
                "repost_count": repost,
                "view_count": views,
            }
        )
    return result


# ---------------------------------------------------------------
# 划分 / 缩放 / 筛选 / 抽样
# ---------------------------------------------------------------

def _split_ids(
    ids: List[str], *, test_ratio: float, random_seed: int
) -> Tuple[List[str], List[str]]:
    if len(ids) < 2:
        raise ValueError("记录总数少于 2 条，无法划分训练/测试分区。")
    rng = np.random.default_rng(random_seed)
    n_test = max(1, int(round(len(ids) * test_ratio)))
    n_test = min(n_test, len(ids) - 1)
    shuffled = rng.permutation(np.asarray(ids)).tolist()
    test_ids = sorted(shuffled[:n_test])
    test_set = set(test_ids)
    train_ids = [sid for sid in ids if sid not in test_set]
    if not train_ids:
        raise ValueError("划分后训练分区为空。")
    return train_ids, test_ids


def _compute_anchor(
    train_records: Sequence[Dict[str, Any]], *, anchor_percentile: float
) -> Tuple[float, List[str]]:
    candidates = [
        r for r in train_records
        if r["repost_count"] > MIN_REPOST_FOR_FILTER and r["view_count"] > MIN_VIEW_FOR_FILTER
    ]
    if not candidates:
        raise ValueError("训练分区中满足转发量>0且浏览量>100的记录为空，无法计算缩放锚点。")
    views = np.asarray([r["view_count"] for r in candidates], dtype=float)
    anchor = float(np.quantile(views, anchor_percentile))
    if anchor <= MIN_VIEW_FOR_FILTER:
        anchor = float(views.max())
    return anchor, [r["story_id"] for r in candidates]


def _apply_scaling(
    records: Sequence[Dict[str, Any]], *, scale_ratio: float, num_agents: int
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for r in records:
        unclipped = int(round(float(r["repost_count"]) * scale_ratio))
        clipped = min(unclipped, num_agents)
        out.append(
            {
                **r,
                "scaled_target": clipped,
                "unclipped_target": unclipped,
                "target_clipped": unclipped > num_agents,
            }
        )
    return out


def _data_num(population_size: int) -> int:
    """设计效应校正后的样本量（与 StoryManager 同款）。"""
    z_score = 1.96
    p = 0.5
    deff = 1.5
    e = 0.05
    n_infinite = ((z_score**2) * p * (1 - p)) / (e**2)
    n_finite = n_infinite / (1 + ((n_infinite - 1) / max(population_size, 1)))
    return int(round(n_finite * deff))


def _sample_stratified(
    records: Sequence[Dict[str, Any]], *, random_seed: int
) -> List[Dict[str, Any]]:
    """十档分层抽样，样本量不超过有效记录数，固定随机种子。"""
    import pandas as pd

    n = len(records)
    sample_size = min(_data_num(n), n)
    if sample_size >= n:
        return list(records)

    df = pd.DataFrame(list(records))
    df["_pct_rank"] = df["scaled_target"].rank(pct=True, method="first")
    df["_bin"] = pd.cut(
        df["_pct_rank"],
        bins=np.linspace(0, 1, 11),
        labels=False,
        include_lowest=True,
    )
    bin_counts = df["_bin"].value_counts().sort_index()
    n_bins = max(len(bin_counts), 1)
    base_per_bin = sample_size // n_bins
    remainder = sample_size % n_bins
    targets = {int(b): base_per_bin for b in bin_counts.index}
    sorted_bins = bin_counts.sort_values(ascending=False).index
    for i in range(remainder):
        targets[int(sorted_bins[i])] += 1

    selected: List[Any] = []
    for bin_label, group in df.groupby("_bin"):
        target_n = targets.get(int(bin_label), base_per_bin)
        take_n = min(len(group), target_n)
        sampled = group.sample(n=take_n, random_state=random_seed) if len(group) > take_n else group
        selected.append(sampled)
    rep = pd.concat(selected, ignore_index=True)
    rep = rep.drop(columns=["_pct_rank", "_bin"])
    return rep.to_dict("records")


# ---------------------------------------------------------------
# 主入口
# ---------------------------------------------------------------

def prepare_dataset(
    source: Path,
    output_dir: Path,
    *,
    num_agents: int,
    text_column: Optional[str],
    retweet_columns: str,
    view_column: str,
    id_column: str,
    test_ratio: float = 0.3,
    random_seed: int = 42,
    anchor_percentile: float = 0.8,
    min_scaled_target: int = 5,
    selection: str = "all",
) -> Path:
    """解析原始观测，产出固定数据包，返回 manifest 路径。"""
    source = Path(source)
    output_dir = Path(output_dir)

    if num_agents <= 0:
        raise ValueError("num_agents 必须大于 0。")
    if not 0 < test_ratio < 1:
        raise ValueError("test_ratio 必须严格在 0 与 1 之间。")
    if not 0 < anchor_percentile <= 1:
        raise ValueError("anchor_percentile 必须在 (0, 1] 区间内。")
    if min_scaled_target < 0:
        raise ValueError("min_scaled_target 必须非负。")
    if selection not in ("all", "stratified"):
        raise ValueError(f"未知选择策略：{selection}")

    if source.suffix.lower() == ".json":
        parsed = _read_json_records(source)
        field_mapping = {
            "id_column": "story_id",
            "text_column": "text",
            "view_column": "view_count",
            "retweet_columns": ["repost_count"],
        }
    else:
        parsed = _read_table_records(
            source,
            text_column=text_column,
            retweet_columns=retweet_columns,
            view_column=view_column,
            id_column=id_column,
        )
        field_mapping = {
            "id_column": id_column,
            "text_column": text_column,
            "view_column": view_column,
            "retweet_columns": [item.strip() for item in retweet_columns.split(",") if item.strip()],
        }

    # 清洗：零浏览量排除并计数
    zero_view_excluded = 0
    cleaned: List[Dict[str, Any]] = []
    seen_ids: set = set()
    for index, record in enumerate(parsed):
        row_no = index + 2
        story_id = record["story_id"]
        if not story_id:
            raise ValueError(f"第 {row_no} 行存在空 ID。")
        if story_id in seen_ids:
            raise ValueError(f"第 {row_no} 行存在重复 ID：{story_id}")
        seen_ids.add(story_id)
        if not record["text"].strip():
            raise ValueError(f"第 {row_no} 行（ID {story_id}）正文为空。")
        if record["view_count"] == 0:
            zero_view_excluded += 1
            continue
        cleaned.append(record)

    if not cleaned:
        raise ValueError("清洗后没有可用记录。")

    cleaned.sort(key=lambda r: r["story_id"])
    all_ids = [r["story_id"] for r in cleaned]
    train_ids, test_ids = _split_ids(all_ids, test_ratio=test_ratio, random_seed=random_seed)

    by_id = {r["story_id"]: r for r in cleaned}
    train_records = [by_id[sid] for sid in train_ids]
    test_records = [by_id[sid] for sid in test_ids]

    anchor, anchor_sample_ids = _compute_anchor(train_records, anchor_percentile=anchor_percentile)
    scale_ratio = num_agents / anchor

    train_scaled = _apply_scaling(train_records, scale_ratio=scale_ratio, num_agents=num_agents)
    test_scaled = _apply_scaling(test_records, scale_ratio=scale_ratio, num_agents=num_agents)

    train_eligible = [
        r for r in train_scaled
        if r["repost_count"] > MIN_REPOST_FOR_FILTER
        and r["view_count"] > MIN_VIEW_FOR_FILTER
        and r["scaled_target"] >= min_scaled_target
    ]
    if not train_eligible:
        raise ValueError("筛选后训练集为空。")
    if selection == "stratified":
        train_eligible = _sample_stratified(train_eligible, random_seed=random_seed)

    train_final = sorted(train_eligible, key=lambda r: r["story_id"])
    test_final = sorted(test_scaled, key=lambda r: r["story_id"])

    def _record(r: Dict[str, Any]) -> Dict[str, Any]:
        return {
            "story_id": r["story_id"],
            "text": r["text"],
            "repost_count": r["repost_count"],
            "view_count": r["view_count"],
            "scaled_target": r["scaled_target"],
            "unclipped_target": r["unclipped_target"],
            "target_clipped": r["target_clipped"],
        }

    train_payload = {
        "schema_version": SCHEMA_VERSION,
        "split": "train",
        "num_agents": num_agents,
        "scale_ratio": scale_ratio,
        "records": [_record(r) for r in train_final],
    }
    test_payload = {
        "schema_version": SCHEMA_VERSION,
        "split": "test",
        "num_agents": num_agents,
        "scale_ratio": scale_ratio,
        "records": [_record(r) for r in test_final],
    }

    with publish_output(output_dir) as temp_dir:
        _write_json(temp_dir / "train.json", train_payload)
        _write_json(temp_dir / "test.json", test_payload)
        train_hash = file_sha256(temp_dir / "train.json")
        test_hash = file_sha256(temp_dir / "test.json")
        manifest = {
            "schema_version": SCHEMA_VERSION,
            "source_path": str(source),
            "source_sha256": file_sha256(source),
            "field_mapping": field_mapping,
            "cleaning_counts": {"zero_view_excluded": zero_view_excluded},
            "split_seed": random_seed,
            "test_ratio": test_ratio,
            "selection": selection,
            "anchor_percentile": anchor_percentile,
            "min_scaled_target": min_scaled_target,
            "num_agents": num_agents,
            "anchor": anchor,
            "anchor_sample_ids": anchor_sample_ids,
            "scale_ratio": scale_ratio,
            "train_ids": [r["story_id"] for r in train_final],
            "test_ids": [r["story_id"] for r in test_final],
            "train_count": len(train_final),
            "test_count": len(test_final),
            "file_hashes": {"train": train_hash, "test": test_hash},
        }
        _write_json(temp_dir / "manifest.json", manifest)

    return output_dir / "manifest.json"


# ---------------------------------------------------------------
# 数据包加载
# ---------------------------------------------------------------

def load_split(
    path: Path, *, expected_split: str
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """读取所请求分区与相邻 manifest，校验版本/字段/散列/ID/规模。

    只读取所请求分区与 manifest，不打开另一分区。
    """
    if expected_split not in ("train", "test"):
        raise ValueError(f"未知分区：{expected_split}")
    path = Path(path)
    manifest_path = path.parent / "manifest.json"
    if not path.exists():
        raise FileNotFoundError(f"找不到分区文件：{path}")
    if not manifest_path.exists():
        raise FileNotFoundError(f"找不到 manifest：{manifest_path}")

    manifest = _read_json(manifest_path)
    if manifest.get("schema_version") != SCHEMA_VERSION:
        raise ValueError(
            f"manifest schema_version 不支持：{manifest.get('schema_version')}"
        )

    partition = _read_json(path)
    if partition.get("schema_version") != SCHEMA_VERSION:
        raise ValueError(f"分区 schema_version 不支持：{partition.get('schema_version')}")
    if partition.get("split") != expected_split:
        raise ValueError(f"分区 split={partition.get('split')!r}，期望 {expected_split!r}")
    if partition.get("num_agents") != manifest.get("num_agents"):
        raise ValueError("分区 num_agents 与 manifest 不一致。")
    if not _almost_equal(partition.get("scale_ratio"), manifest.get("scale_ratio")):
        raise ValueError("分区 scale_ratio 与 manifest 不一致。")

    records = partition.get("records")
    if not isinstance(records, list):
        raise ValueError("分区 records 必须是数组。")
    for index, record in enumerate(records):
        if not isinstance(record, dict):
            raise ValueError(f"records[{index}] 必须是 JSON 对象。")
        for field in _REQUIRED_RECORD_FIELDS:
            if field not in record:
                raise ValueError(f"records[{index}] 缺少字段：{field}")

    partition_ids = sorted(str(r["story_id"]) for r in records)
    manifest_ids = sorted(str(sid) for sid in manifest.get(f"{expected_split}_ids", []))
    if partition_ids != manifest_ids:
        raise ValueError("分区记录 ID 与 manifest 不一致。")
    expected_count = manifest.get(f"{expected_split}_count")
    if expected_count is not None and len(records) != int(expected_count):
        raise ValueError("分区记录数与 manifest 不一致。")

    file_hashes = manifest.get("file_hashes") or {}
    expected_hash = file_hashes.get(expected_split)
    if not expected_hash:
        raise ValueError(f"manifest 缺少 {expected_split} 分区散列。")
    if file_sha256(path) != expected_hash:
        raise ValueError(f"分区文件散列不匹配：{path}")

    return partition, manifest


def _almost_equal(a: Any, b: Any) -> bool:
    if a is None or b is None:
        return a == b
    try:
        return abs(float(a) - float(b)) < 1e-9
    except (TypeError, ValueError):
        return a == b
