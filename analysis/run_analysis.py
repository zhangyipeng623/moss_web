from __future__ import annotations

import argparse
import asyncio
import csv
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Awaitable, Callable
from zoneinfo import ZoneInfo

import numpy as np


PORTRAIT_USER_SUPPORTED_COLUMNS = (
    "用户名",
    "昵称",
    "简介",
    "性别",
    "地域",
    "关注",
    "粉丝",
    "收藏",
    "源用户名",
    "用户地址",
    "创建时间戳",
    "头像链接",
)
PORTRAIT_USER_REQUIRED_COLUMNS = (
    "用户名",
    "昵称",
    "简介",
    "性别",
    "地域",
    "关注",
    "粉丝",
    "收藏",
    "创建时间戳",
)
PORTRAIT_POST_SUPPORTED_COLUMNS = (
    "用户名",
    "发文内容",
    "发布时间",
    "发布时间戳",
    "发文类型",
    "点赞数",
    "评论数",
    "转发数",
    "帖子ID",
    "推文ID",
    "post_id",
    "id",
)
PORTRAIT_POST_REQUIRED_COLUMNS = (
    "用户名",
    "发文内容",
    "发布时间",
    "发布时间戳",
    "发文类型",
)


def parse_args() -> argparse.Namespace:
    """解析命令行参数。"""
    parser = argparse.ArgumentParser(description="离线分析脚本")
    subparsers = parser.add_subparsers(dest="command", required=True)

    portrait_parser = subparsers.add_parser("portrait", help="生成用户画像")
    portrait_parser.add_argument(
        "--data-path",
        help="数据目录路径，目录下默认包含 user.xlsx 和 post.xlsx",
    )
    portrait_parser.add_argument(
        "--user-name",
        help="目标用户名，单用户模式必填；使用 --batch 批量模式时无需提供",
    )
    portrait_parser.add_argument(
        "--user-file",
        default="user.xlsx",
        help="用户信息文件名，默认 user.xlsx",
    )
    portrait_parser.add_argument(
        "--post-file",
        default="post.xlsx",
        help="帖子文件名，默认 post.xlsx",
    )
    portrait_parser.add_argument(
        "--output",
        help=(
            "单用户模式：画像输出 JSON 路径，默认 analysis_outputs/portraits/<user_name>.json；"
            "批量模式：画像输出目录，默认 analysis_outputs/portraits/"
        ),
    )
    portrait_parser.add_argument(
        "--portraits-dir",
        help=(
            "画像输出目录（推荐，便于按数据集分目录）："
            "单用户模式写入 <portraits-dir>/<user_name>.json；批量模式直接作为输出目录。"
            "默认 analysis_outputs/portraits/。优先级低于 --output。"
        ),
    )
    portrait_parser.add_argument(
        "--json-schema",
        action=argparse.BooleanOptionalAction,
        default=True,
        help=(
            "启用 JSON Schema 结构化输出（OpenRouter response_format=json_schema），"
            "强制模型返回固定格式。模型/端点不支持时自动回退纯文本解析。"
            "用 --no-json-schema 关闭。"
        ),
    )
    portrait_parser.add_argument(
        "--reference-time",
        help=(
            "画像生成参考时间，未提供时会交互输入。"
            "若不带时区，默认按 Asia/Shanghai 解析，例如 2026-04-01 12:00:00"
        ),
    )
    portrait_parser.add_argument(
        "--model",
        default=os.environ.get("MODEL", "gpt-4o"),
        help="模型名称，默认读取环境变量 MODEL",
    )
    portrait_parser.add_argument("--timeout", type=int, default=180, help="模型调用超时时间")
    portrait_parser.add_argument(
        "--api-key-env",
        default="API_KEY",
        help="读取模型密钥的环境变量名",
    )
    portrait_parser.add_argument(
        "--base-url-env",
        default="BASE_URL",
        help="读取模型服务地址的环境变量名",
    )
    portrait_parser.add_argument(
        "--batch",
        action="store_true",
        help="批量模式：自动读取 user.xlsx 中全部用户并并行生成画像，此时 --user-name 无需提供",
    )
    portrait_parser.add_argument(
        "--concurrency",
        type=int,
        default=10,
        help="批量模式下同时生成的用户画像数量（并发数），默认 10",
    )

    retier_parser = subparsers.add_parser("retier", help="对已有画像 JSON 重新按影响力分级")
    retier_parser.add_argument(
        "--portraits-dir",
        required=True,
        help="存放画像 JSON 文件的目录路径",
    )

    recommender_parser = subparsers.add_parser("recommender", help="反推推荐参数（全量公共概率校准）")
    recommender_parser.add_argument(
        "--train-file",
        help="训练分区 train.json 路径（由 scripts/prepare_recommender_data.py 生成）",
    )
    recommender_parser.add_argument(
        "--portraits-dir",
        help="用户画像 JSON 目录路径（由 portrait --batch 生成），用于语义处理和种群扩增",
    )
    recommender_parser.add_argument(
        "--output-dir",
        help="输出目录（写入 model.json 与 calibration_profile.yaml，已存在则报错）",
    )
    recommender_parser.add_argument(
        "--embedding-model",
        default="Alibaba-NLP/gte-multilingual-base",
        help="文本嵌入模型名称，默认 Alibaba-NLP/gte-multilingual-base",
    )
    recommender_parser.add_argument(
        "--n-cpu",
        type=int,
        default=4,
        help="并行校准使用的 CPU 核心数，默认 4",
    )
    recommender_parser.add_argument(
        "--random-seed",
        type=int,
        default=42,
        help="全链路随机种子（种群合成/引擎/校准共用），默认 42",
    )
    recommender_parser.add_argument(
        "--time-scale",
        type=float,
        default=3600.0,
        help="实验 system_time.time_scale（每步秒数），默认 3600",
    )
    recommender_parser.add_argument(
        "--rounds",
        type=int,
        default=24,
        help="实验 runtime.rounds（总步数，即模拟 duration），默认 24",
    )
    recommender_parser.add_argument(
        "--max-iterations",
        type=int,
        default=3,
        help="公共概率与权重交替校准最大迭代次数，默认 3",
    )
    recommender_parser.add_argument(
        "--n-repeats",
        type=int,
        default=5,
        help="每条内容每次评估的重复次数，默认 5",
    )
    recommender_parser.add_argument(
        "--p-trials",
        type=int,
        default=20,
        help="公共概率搜索 trial 数，默认 20",
    )
    recommender_parser.add_argument(
        "--weight-trials",
        type=int,
        default=50,
        help="权重搜索 trial 数，默认 50",
    )
    # 旧参数：保留解析以产生定向迁移错误，不接受训练
    recommender_parser.add_argument("--data-file", default=None, help=argparse.SUPPRESS)
    recommender_parser.add_argument("--input", default=None, help=argparse.SUPPRESS)
    recommender_parser.add_argument("--retweet-columns", default=None, help=argparse.SUPPRESS)
    recommender_parser.add_argument("--view-column", default=None, help=argparse.SUPPRESS)
    recommender_parser.add_argument("--id-column", default=None, help=argparse.SUPPRESS)
    recommender_parser.add_argument("--anchor-percentile", type=float, default=None, help=argparse.SUPPRESS)
    recommender_parser.add_argument("--min-scaled-target", type=int, default=None, help=argparse.SUPPRESS)
    recommender_parser.add_argument("--num-agents", type=int, default=None, help=argparse.SUPPRESS)
    recommender_parser.add_argument("--robustness-seeds", default=None, help=argparse.SUPPRESS)
    recommender_parser.add_argument("--output", default=None, help=argparse.SUPPRESS)

    return parser.parse_args()


async def run_portrait_command(args: argparse.Namespace) -> None:
    """执行用户画像生成命令。"""
    from langchain_openai import ChatOpenAI

    from analysis.user_portrait_generator import PortraitGenerationError, UserPortraitGenerator

    source, global_event = _resolve_portrait_source(args)
    reference_timestamp, reference_time_text = _resolve_portrait_reference_time(args.reference_time)
    default_portrait_path = (
        f"{args.portraits_dir.rstrip('/')}/{source.user_name}.json"
        if args.portraits_dir
        else f"analysis_outputs/portraits/{source.user_name}.json"
    )
    output_path = _resolve_output_path(
        args.output,
        default_relative_path=default_portrait_path,
    )

    api_key = os.environ.get(args.api_key_env)
    if not api_key:
        raise ValueError(f"环境变量 {args.api_key_env} 未设置，无法调用画像生成模型。")

    llm = ChatOpenAI(
        model=args.model,
        api_key=api_key,
        base_url=os.environ.get(args.base_url_env),
        timeout=args.timeout,
    )
    generator = UserPortraitGenerator(reference_timestamp=reference_timestamp)
    llm_callable = _build_llm_callable(llm, use_json_schema=args.json_schema)

    last_error = ""
    max_attempts = 3
    for attempt in range(1, max_attempts + 1):
        try:
            profile = await generator.generate_portrait(
                source=source,
                llm_callable=llm_callable,
                global_event=global_event,
            )
            profile["influence_tier"] = 4
            profile["influence_tier_label"] = "Early Adopters"
            _write_json(output_path, profile)
            print(
                f"用户画像已写入：{output_path}；参考时间：{reference_time_text}；"
                f"尝试次数：{attempt}"
            )
            return
        except PortraitGenerationError as exc:
            last_error = str(exc)
        except Exception as exc:  # noqa: BLE001
            last_error = str(exc)

        if attempt < max_attempts:
            print(
                f"用户 {source.user_name} 第 {attempt} 次画像生成失败，准备重试：{last_error}"
            )

    failure_report_path = _record_failed_portrait_user(
        output_path=output_path,
        user_name=source.user_name,
        reference_time_text=reference_time_text,
        error_message=last_error or "未知错误",
    )
    raise RuntimeError(
        f"用户 {source.user_name} 连续 {max_attempts} 次画像生成失败，"
        f"未输出画像文件。失败名单已写入：{failure_report_path}"
    )


async def run_portrait_batch_command(args: argparse.Namespace) -> None:
    """批量生成全部用户画像，并按 P84 阈值自动分 L4/L5（Rogers 5 级框架）。"""
    from langchain_openai import ChatOpenAI

    from analysis.user_portrait_generator import PortraitGenerationError, UserPortraitGenerator

    data_path = Path(args.data_path).resolve()
    all_usernames = _extract_all_usernames(data_path, args.user_file)

    reference_timestamp, reference_time_text = _resolve_portrait_reference_time(
        args.reference_time
    )

    output_dir = (
        Path(args.output).resolve()
        if args.output
        else (
            Path(args.portraits_dir).resolve()
            if args.portraits_dir
            else Path("analysis_outputs/portraits").resolve()
        )
    )
    output_dir.mkdir(parents=True, exist_ok=True)

    api_key = os.environ.get(args.api_key_env)
    if not api_key:
        raise ValueError(f"环境变量 {args.api_key_env} 未设置，无法调用画像生成模型。")

    generator = UserPortraitGenerator(reference_timestamp=reference_timestamp)

    # ----------------------------------------------------------------
    # Phase 1: 预计算所有用户的 account_influence（不调 LLM）
    # 问题 3 修复：两张表只读一次，循环内按用户名查索引，
    # 不再为每个用户重复整表读取（原为 O(N × 表大小)）。
    # ----------------------------------------------------------------
    user_rows = _load_tabular_rows(data_path / args.user_file)
    post_rows = _load_tabular_rows(data_path / args.post_file)
    _validate_tabular_columns(
        rows=user_rows,
        path=data_path / args.user_file,
        required_columns=PORTRAIT_USER_REQUIRED_COLUMNS,
        supported_columns=PORTRAIT_USER_SUPPORTED_COLUMNS,
    )
    _validate_tabular_columns(
        rows=post_rows,
        path=data_path / args.post_file,
        required_columns=PORTRAIT_POST_REQUIRED_COLUMNS,
        supported_columns=PORTRAIT_POST_SUPPORTED_COLUMNS,
    )

    # B-1（R4）：一次建索引，循环内 O(1) 查表（I/O 已提外 + CPU 侧不再逐用户扫全表）
    users_by_name, posts_by_user = _build_portrait_user_indexes(
        user_rows=user_rows,
        post_rows=post_rows,
    )

    user_records: list[dict[str, Any]] = []
    phase1_failed: list[dict[str, str]] = []

    print(f"阶段 1/2：预计算影响力分数，共 {len(all_usernames)} 个用户 ...\n")
    for username in all_usernames:
        try:
            target_name = _normalize_username(username)
            user_row = _select_best_user_row(users_by_name.get(target_name, []))
            if user_row is None:
                raise ValueError(f"在用户表中找不到用户：{username}")
            source, _ = _build_portrait_source_from_matched(
                user_row=user_row,
                matched_posts=posts_by_user.get(target_name, []),
                user_name=username,
            )
        except Exception as exc:
            phase1_failed.append(
                {
                    "user_name": username,
                    "reference_time": reference_time_text,
                    "error_message": str(exc),
                }
            )
            continue
        stats = generator.analyze_stats(source)
        user_records.append(
            {
                "username": username,
                "canonical_name": source.user_name,
                "source": source,
                "account_influence": stats.account_influence,
                "fans_count": stats.fans_count,
                "post_count": stats.post_count,
            }
        )

    if not user_records:
        raise RuntimeError("没有可用用户数据，无法进行批量画像生成。")

    # ----------------------------------------------------------------
    # Phase 2: P84 单阈值切分 L4/L5（Rogers 5 级框架——仅含 Early Adopters + Innovators）
    # ----------------------------------------------------------------
    sorted_records = sorted(user_records, key=lambda r: r["account_influence"])
    n = len(sorted_records)
    # 问题 4 修复：P84 用 np.percentile 线性插值，替代整数索引近似
    if n >= 5:
        scores = np.asarray(
            [record["account_influence"] for record in sorted_records],
            dtype=float,
        )
        p84_threshold = float(np.percentile(scores, 84))
        if n < 30:
            print(f"提示：样本量 n={n} 较小，P84 分级阈值可能不稳定。")
    else:
        p84_threshold = float("inf")

    tier_counts: dict[int, int] = {4: 0, 5: 0}
    for record in sorted_records:
        score = record["account_influence"]
        if score > p84_threshold and p84_threshold != float("inf"):
            record["tier"] = 5
        else:
            record["tier"] = 4
        tier_counts[record["tier"]] += 1
        record["tier_label"] = _tier_label(record["tier"])

    print(f"影响力分级完成（阈值 P84={p84_threshold:.2f}）："
          f"L5 Innovators {tier_counts[5]} 人 / L4 Early Adopters {tier_counts[4]} 人\n")

    # ----------------------------------------------------------------
    # Phase 3: 调 LLM 生成画像，注入 influence_tier（跳过已存在文件）
    # ----------------------------------------------------------------
    llm = ChatOpenAI(
        model=args.model,
        api_key=api_key,
        base_url=os.environ.get(args.base_url_env),
        timeout=args.timeout,
    )
    llm_callable = _build_llm_callable(llm, use_json_schema=args.json_schema)

    existing_files = {
        p.stem for p in output_dir.glob("*.json") if p.name != "failed_users.json"
    }
    pending_records = [
        r for r in user_records if r["canonical_name"] not in existing_files
    ]
    skipped_count = len(user_records) - len(pending_records)

    if skipped_count > 0:
        print(f"跳过 {skipped_count} 个已有画像的用户")
        # 对已跳过的用户原地更新 tier 字段
        for record in user_records:
            if record["canonical_name"] in existing_files:
                existing_path = output_dir / f"{record['canonical_name']}.json"
                try:
                    existing_profile = _load_json_object(existing_path)
                except Exception:
                    continue
                existing_profile["influence_tier"] = record["tier"]
                existing_profile["influence_tier_label"] = record["tier_label"]
                _write_json(existing_path, existing_profile)
        print()

    if not pending_records:
        print("所有用户画像均已存在，无需生成。")
        return

    success_count = 0
    phase3_failed: list[dict[str, str]] = []
    total = len(pending_records)
    concurrency = max(1, int(args.concurrency))

    print(f"阶段 2/2：批量生成用户画像（{total} 个待生成，并发 {concurrency}）...\n")

    semaphore = asyncio.Semaphore(concurrency)

    async def _generate_one(record: dict[str, Any]) -> dict[str, Any]:
        """并发生成单个用户画像（带 3 次重试），返回结构化结果。"""
        async with semaphore:
            username = record["username"]
            tier = record["tier"]
            last_error = ""
            max_attempts = 3
            profile = None
            try:
                for attempt in range(1, max_attempts + 1):
                    try:
                        profile = await generator.generate_portrait(
                            source=record["source"],
                            llm_callable=llm_callable,
                        )
                        break
                    except PortraitGenerationError as exc:
                        last_error = str(exc)
                    except Exception as exc:  # noqa: BLE001
                        last_error = str(exc)
                    if attempt < max_attempts:
                        print(
                            f"    {username} 第 {attempt} 次失败，重试中 ...",
                            flush=True,
                        )

                if profile is None:
                    return {
                        "username": username,
                        "success": False,
                        "error": last_error or "未知错误",
                    }

                profile["influence_tier"] = tier
                profile["influence_tier_label"] = record["tier_label"]
                output_path = output_dir / f"{record['canonical_name']}.json"
                _write_json(output_path, profile)
                return {"username": username, "success": True, "error": ""}
            except Exception as exc:  # noqa: BLE001
                return {"username": username, "success": False, "error": str(exc)}

    worker_tasks = [
        asyncio.create_task(_generate_one(record)) for record in pending_records
    ]
    done_count = 0
    for completed in asyncio.as_completed(worker_tasks):
        result = await completed
        done_count += 1
        if result["success"]:
            success_count += 1
            print(f"[{done_count}/{total}] {result['username']} 完成")
        else:
            phase3_failed.append(
                {
                    "user_name": result["username"],
                    "reference_time": reference_time_text,
                    "error_message": result["error"],
                }
            )
            print(f"[{done_count}/{total}] {result['username']} 失败：{result['error']}")

    # ----------------------------------------------------------------
    # 汇总
    # ----------------------------------------------------------------
    full_total = len(user_records)
    all_failed = phase1_failed + phase3_failed
    print(
        f"\n批量画像生成结束：本次生成 {success_count}/{total}；"
        f"总计 {full_total} 人（已跳过 {skipped_count}），失败 {len(all_failed)} 人"
    )
    print(f"分级统计：L5 Innovators {tier_counts[5]} 人 / L4 Early Adopters {tier_counts[4]} 人")

    if all_failed:
        failure_report_path = output_dir / "failed_users.json"
        payload: dict[str, Any] = {
            "failed_users": [item["user_name"] for item in all_failed],
            "failed_details": all_failed,
        }
        _write_json(failure_report_path, payload)
        print(f"失败名单已写入：{failure_report_path}")


def _tier_label(tier: int) -> str:
    """Rogers 5 级创新扩散标签。"""
    labels = {
        1: "Laggards",
        2: "Late Majority",
        3: "Early Majority",
        4: "Early Adopters",
        5: "Innovators",
    }
    return labels.get(tier, f"Lv{tier}")


def run_retier_command(args: argparse.Namespace) -> None:
    """对已有画像 JSON 目录重新计算影响力分级，并回写每个文件。"""
    portraits_dir = Path(args.portraits_dir).resolve()
    if not portraits_dir.is_dir():
        raise ValueError(f"目录不存在：{portraits_dir}")

    json_files = sorted(portraits_dir.glob("*.json"))
    if not json_files:
        raise ValueError(f"目录中没有 JSON 文件：{portraits_dir}")

    # 读取所有画像，提取 account_influence
    records: list[dict[str, Any]] = []
    for path in json_files:
        try:
            profile = _load_json_object(path)
        except Exception:
            print(f"跳过无法解析的文件：{path.name}")
            continue
        stats = profile.get("stats")
        if not isinstance(stats, dict):
            print(f"跳过缺少 stats 字段的文件：{path.name}")
            continue
        score = stats.get("account_influence")
        if not isinstance(score, (int, float)):
            print(f"跳过缺少 account_influence 的文件：{path.name}")
            continue
        records.append(
            {
                "path": path,
                "profile": profile,
                "account_influence": float(score),
                "user_id": profile.get("user_id", path.stem),
            }
        )

    if not records:
        raise RuntimeError("没有可用的画像文件（均缺少 stats.account_influence）。")

    # 按 account_influence 排序，计算 P84 单阈值
    sorted_records = sorted(records, key=lambda r: r["account_influence"])
    n = len(sorted_records)
    # 问题 4 修复：P84 用 np.percentile 线性插值，替代整数索引近似
    if n >= 5:
        scores = np.asarray(
            [record["account_influence"] for record in sorted_records],
            dtype=float,
        )
        p84_threshold = float(np.percentile(scores, 84))
        if n < 30:
            print(f"提示：样本量 n={n} 较小，P84 分级阈值可能不稳定。")
    else:
        p84_threshold = float("inf")

    tier_counts: dict[int, int] = {4: 0, 5: 0}
    for record in sorted_records:
        score = record["account_influence"]
        if score > p84_threshold and p84_threshold != float("inf"):
            tier = 5
        else:
            tier = 4
        record["tier"] = tier
        tier_counts[tier] += 1

    # 回写每个 JSON
    for record in sorted_records:
        record["profile"]["influence_tier"] = record["tier"]
        record["profile"]["influence_tier_label"] = _tier_label(record["tier"])
        _write_json(record["path"], record["profile"])

    print(
        f"retier 完成，共处理 {n} 个画像文件\n"
        f"阈值 P84={p84_threshold:.2f}\n"
        f"L5 Innovators {tier_counts[5]} 人 / L4 Early Adopters {tier_counts[4]} 人"
    )


def _generate_calibration_yaml(
    output_dir: str,
    train_file: str,
    portraits_dir: str,
    best_weights: dict,
    p_base_global: float | None,
    fit_diagnostics: dict,
    embedding_model: str,
    num_seed_users: int,
    abm_population_size: int,
    time_scale: float = 3600.0,
    rounds: int = 24,
) -> str:
    """在 output_dir 下生成 calibration_profile.yaml（沿用线上字段，新增可选 p_base_global）。

    D-1：时间基准显式写入——system_time.time_scale、runtime.rounds。
    """
    import yaml

    from core.calibration_profile import (
        CalibrationProfile,
        EmbeddingConfig,
        ExperimentYamlConfig,
        MetaInfo,
        RecommenderConfig,
        RecommenderWeights,
    )
    from core.experiment_config import RuntimeExperimentConfig, SystemTimeExperimentConfig

    profile = CalibrationProfile(
        experiment=ExperimentYamlConfig(
            system_time=SystemTimeExperimentConfig(
                start_time="2026-05-17T12:00:00",
                time_scale=time_scale,
            ),
            runtime=RuntimeExperimentConfig(rounds=rounds),
        ),
        meta=MetaInfo(
            generated_at=datetime.now().isoformat(),
            portraits_dir=str(Path(portraits_dir).resolve()),
            num_seed_users=num_seed_users,
            abm_population_size=abm_population_size,
            input_data_file=str(Path(train_file).name),
        ),
        recommender=RecommenderConfig(
            weights=RecommenderWeights(
                w_interest=best_weights.get("w_i", 0.35),
                w_popularity=best_weights.get("w_pop", 0.25),
                w_time=best_weights.get("w_time", 0.25),
                w_random=best_weights.get("w_rand", 0.15),
            ),
            decay_lambda=best_weights.get("decay_lambda", 0.5),
            calibrated_p_base={},
            p_base_global=p_base_global,
            fit_diagnostics=fit_diagnostics or {},
        ),
        embedding=EmbeddingConfig(model_name=embedding_model),
    )

    yaml_path = Path(output_dir) / "calibration_profile.yaml"
    yaml_content = yaml.safe_dump(
        profile.model_dump(exclude_none=True),
        allow_unicode=True,
        default_flow_style=False,
        sort_keys=False,
    )

    header = (
        f"# calibration_profile.yaml\n"
        f"# 自动生成于：{profile.meta.generated_at}（recommender 命令输出）\n"
        f"# 用途：唯一的模拟配置文件，包含离线校准结果 + 在线模拟全部参数\n"
        f"# 位置：{yaml_path}\n"
        f"# 使用：python main.py --config {yaml_path}\n"
        f"#\n"
        f"# 大小模型分层（可选）：在 experiment 段下补一段 llm_small 即可启用，\n"
        f"# simple（mass 普通用户）将改用小模型；不写 llm_small 则全员回退大模型（对照组）：\n"
        f"#   experiment:\n"
        f"#     llm_small:\n"
        f"#       model: Qwen2.5-7B-Instruct\n"
        f"#       timeout: 180\n"
        f"#       api_key_env: API_KEY_SMALL\n"
        f"#       base_url_env: BASE_URL_SMALL\n"
        f"#\n"
        f"# L1-L3 大众用户动态抽取（可选）：simulation.l1_l3_pool.enabled=true 时，\n"
        f"# 每次 main.py 启动按 Rogers 比例（16/34/34，以 L4+L5 数量为锚）从候选池 csv 随机抽取 L1-L3 simple 用户：\n"
        f"#   simulation:\n"
        f"#     l1_l3_pool:\n"
        f"#       enabled: true\n"
        f"#       csv_path: data/l1_l3_pool/users_all_fields_deduped.csv\n"
        f"#       exclude_verified: true\n"
        f"#       seed: 42\n"
        f"\n"
    )
    with open(yaml_path, "w", encoding="utf-8") as f:
        f.write(header + yaml_content)

    return str(yaml_path)


def _raise_if_legacy_recommender_args(args: argparse.Namespace) -> None:
    """检测旧 recommender 参数并给出定向迁移错误。"""
    legacy = {
        "--data-file": getattr(args, "data_file", None),
        "--input": getattr(args, "input", None),
        "--retweet-columns": getattr(args, "retweet_columns", None),
        "--view-column": getattr(args, "view_column", None),
        "--id-column": getattr(args, "id_column", None),
        "--anchor-percentile": getattr(args, "anchor_percentile", None),
        "--min-scaled-target": getattr(args, "min_scaled_target", None),
        "--num-agents": getattr(args, "num_agents", None),
        "--robustness-seeds": getattr(args, "robustness_seeds", None),
        "--output": getattr(args, "output", None),
    }
    provided = [name for name, value in legacy.items() if value is not None]
    if provided:
        joined = "、".join(provided)
        raise ValueError(
            f"旧 recommender 参数（{joined}）已迁移：请先用 "
            f"scripts/prepare_recommender_data.py 准备数据包，再用 --train-file 训练。"
        )


def run_recommender_command(args: argparse.Namespace) -> None:
    """执行推荐参数反推命令（公共概率全量校准 + 可重建模型）。"""
    from analysis.recommender_data import file_sha256, load_split, publish_output
    from analysis.recommender_parameter_inference import (
        RecommendationParameterInferer,
        load_portrait_bundle,
    )

    _raise_if_legacy_recommender_args(args)

    if not args.train_file:
        raise ValueError("请提供 --train-file（训练分区 train.json 路径）。")
    if not args.portraits_dir:
        raise ValueError("请提供 --portraits-dir（用户画像目录）。")
    if not args.output_dir:
        raise ValueError("请提供 --output-dir（输出目录）。")

    train_path = Path(args.train_file).resolve()
    train_partition, manifest = load_split(train_path, expected_split="train")

    portraits_dir = Path(args.portraits_dir).resolve()
    if not portraits_dir.is_dir():
        raise ValueError(f"画像目录不存在：{portraits_dir}")
    personas, portrait_manifest = load_portrait_bundle(portraits_dir)
    if not personas:
        raise ValueError(f"画像目录没有可用画像：{portraits_dir}")

    num_agents = int(train_partition["num_agents"])
    records = train_partition["records"]
    if not records:
        raise ValueError("训练分区 records 为空。")

    print(
        f"ABM 规模: {num_agents} 人（种子 {len(personas)} 个用户画像，训练 {len(records)} 条）"
    )

    inferer = RecommendationParameterInferer(
        num_agents=num_agents,
        min_scaled_target=int(manifest.get("min_scaled_target", 5)),
        embedding_model=args.embedding_model,
        n_cpu=args.n_cpu,
        target_size_for_sampling=num_agents,
        random_seed=args.random_seed,
        time_scale=args.time_scale,
    )
    inferer.load_portraits(personas)
    inferer.load_prepared_stories(records)
    inferer.precompute_interests()

    result = inferer.run_global_calibration(
        iterations=args.max_iterations,
        duration=args.rounds,
        n_repeats=args.n_repeats,
        p_trials=args.p_trials,
        weight_trials=args.weight_trials,
    )

    env = inferer.environment_snapshot()
    weights = result["weights"]
    p_base_global = result["p_base_global"]
    diagnostics = result["diagnostics"]

    model = {
        "schema_version": 1,
        "data": {
            "train_hash": file_sha256(train_path),
            "manifest_hash": file_sha256(train_path.parent / "manifest.json"),
            "train_story_ids": [str(r["story_id"]) for r in records],
            "scale_ratio": train_partition["scale_ratio"],
        },
        "weights": {k: weights[k] for k in ("w_i", "w_pop", "w_time", "w_rand")},
        "decay_lambda": weights["decay_lambda"],
        "p_base_mode": "global",
        "p_base_global": p_base_global,
        "training": {
            "budget": {
                "iterations": args.max_iterations,
                "p_trials": args.p_trials,
                "weight_trials": args.weight_trials,
                "n_repeats": args.n_repeats,
                "duration": args.rounds,
            },
            "loss_name": diagnostics["loss_name"],
            "best_iteration": diagnostics["best_iteration"],
            "best_loss": diagnostics["best_loss"],
            "replay_loss": diagnostics["replay_loss"],
            "rounds": diagnostics["rounds"],
        },
        "environment": {
            "training_seed": args.random_seed,
            **env,
            "rounds": args.rounds,
        },
        "portraits": {
            "num_seed_users": len(personas),
            "files": portrait_manifest,
        },
        "embedding": {
            "model_name": args.embedding_model,
            "normalize_embeddings": True,
            "versions": _collect_env_versions(),
        },
    }

    output_dir = Path(args.output_dir).resolve()
    with publish_output(output_dir) as temp_dir:
        _write_json(temp_dir / "model.json", model)
        _generate_calibration_yaml(
            output_dir=str(temp_dir),
            train_file=str(train_path),
            portraits_dir=str(portraits_dir),
            best_weights=weights,
            p_base_global=p_base_global,
            fit_diagnostics=diagnostics,
            embedding_model=args.embedding_model,
            num_seed_users=len(personas),
            abm_population_size=num_agents,
            time_scale=args.time_scale,
            rounds=args.rounds,
        )

    print(f"训练模型已写入：{output_dir / 'model.json'}")
    print(f"配置文件已写入：{output_dir / 'calibration_profile.yaml'}")


def _resolve_portrait_source(
    args: argparse.Namespace,
) -> tuple[UserProfileSource, str]:
    """解析画像命令输入来源。"""
    if not args.data_path:
        raise ValueError("请提供 --data-path。")
    if not args.user_name:
        raise ValueError("使用 Excel 数据目录模式时必须提供 --user-name。")

    return _build_portrait_source_from_excel(
        data_path=Path(args.data_path).resolve(),
        user_name=args.user_name,
        user_file=args.user_file,
        post_file=args.post_file,
    )


def _build_portrait_source_from_excel(
    data_path: Path,
    user_name: str,
    user_file: str,
    post_file: str,
) -> tuple[UserProfileSource, str]:
    """按原有 Excel 数据结构提取单个用户的资料和帖子（单用户模式入口）。"""
    user_path = data_path / user_file
    post_path = data_path / post_file
    user_rows = _load_tabular_rows(user_path)
    post_rows = _load_tabular_rows(post_path)
    _validate_tabular_columns(
        rows=user_rows,
        path=user_path,
        required_columns=PORTRAIT_USER_REQUIRED_COLUMNS,
        supported_columns=PORTRAIT_USER_SUPPORTED_COLUMNS,
    )
    _validate_tabular_columns(
        rows=post_rows,
        path=post_path,
        required_columns=PORTRAIT_POST_REQUIRED_COLUMNS,
        supported_columns=PORTRAIT_POST_SUPPORTED_COLUMNS,
    )
    return _build_portrait_source_from_rows(
        user_rows=user_rows,
        post_rows=post_rows,
        user_name=user_name,
    )


def _build_portrait_source_from_rows(
    user_rows: list[dict[str, Any]],
    post_rows: list[dict[str, Any]],
    user_name: str,
) -> tuple[UserProfileSource, str]:
    """从已读入内存的用户表/帖子表构建单用户画像来源（线性匹配版本）。

    批量模式请先用 _build_portrait_user_indexes 建索引，再调用
    _build_portrait_source_from_matched（B-1 / R4：循环内 O(1) 查表）。
    """
    target_name = _normalize_username(user_name)
    if not target_name:
        raise ValueError("user_name 不能为空。")

    matched_user_rows = [
        row
        for row in user_rows
        if _normalize_username(row.get("用户名")) == target_name
    ]
    user_row = _select_best_user_row(matched_user_rows)
    if user_row is None:
        raise ValueError(f"在用户表中找不到用户：{user_name}")

    matched_posts = [
        row
        for row in post_rows
        if _normalize_username(row.get("用户名")) == target_name
    ]
    return _build_portrait_source_from_matched(
        user_row=user_row,
        matched_posts=matched_posts,
        user_name=user_name,
    )


def _build_portrait_user_indexes(
    user_rows: list[dict[str, Any]],
    post_rows: list[dict[str, Any]],
) -> tuple[dict[str, list[dict[str, Any]]], dict[str, list[dict[str, Any]]]]:
    """按规范化用户名对用户表/帖子表各建一次索引（B-1 / R4）。

    返回 (users_by_name, posts_by_user)，批量循环内 O(1) 查表，
    替代逐用户线性扫全表 + 逐行 normalize 的 O(N × 表大小) CPU 开销。
    """
    from collections import defaultdict

    users_by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in user_rows:
        name = _normalize_username(row.get("用户名"))
        if name:
            users_by_name[name].append(row)

    posts_by_user: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in post_rows:
        name = _normalize_username(row.get("用户名"))
        if name:
            posts_by_user[name].append(row)
    return dict(users_by_name), dict(posts_by_user)


def _build_portrait_source_from_matched(
    user_row: dict[str, Any],
    matched_posts: list[dict[str, Any]],
    user_name: str,
) -> tuple[UserProfileSource, str]:
    """从已匹配好的用户行与帖子子集构建画像来源（批量索引版核心）。"""
    from analysis.user_portrait_generator import UserPost, UserProfileSource

    canonical_user_name = _normalize_scalar(user_row.get("用户名")) or user_name
    if not matched_posts:
        raise ValueError(f"在帖子表中找不到用户 {user_name} 的帖子。")

    ordered_posts = sorted(
        matched_posts,
        key=lambda row: (
            _resolve_post_timestamp(row),
            _normalize_scalar(row.get("发布时间")),
            _safe_float(row.get("转发数", 0))
            + _safe_float(row.get("评论数", 0))
            + _safe_float(row.get("点赞数", 0)),
        ),
        reverse=True,
    )

    user_info = {
        "username": canonical_user_name,
        "nickname": _normalize_scalar(user_row.get("昵称")),
        "description": _normalize_scalar(user_row.get("简介")),
        "sex": _normalize_scalar(user_row.get("性别")),
        "region": _normalize_scalar(user_row.get("地域")),
        "follow_count": _safe_int(user_row.get("关注", 0)),
        "fans_count": _safe_int(user_row.get("粉丝", 0)),
        "collect_count": _safe_int(user_row.get("收藏", 0)),
        "follow_user": _normalize_scalar(user_row.get("源用户名")),
        "home_page_url": _normalize_scalar(user_row.get("用户地址")),
        "register_timestamp": _safe_int(user_row.get("创建时间戳", 0)),
        "profile_img_url": _normalize_scalar(user_row.get("头像链接")),
    }

    posts: list[UserPost] = []
    seen_post_keys: set[tuple[Any, ...]] = set()
    for row in ordered_posts:
        content = _normalize_scalar(row.get("发文内容"))
        if not content:
            continue
        timestamp = _resolve_post_timestamp(row)
        publish_time = _normalize_scalar(row.get("发布时间"))
        content_type = _normalize_scalar(row.get("发文类型")) or "Social Post"
        source_post_id = _safe_optional_int(
            row.get("帖子ID")
            or row.get("post_id")
            or row.get("推文ID")
            or row.get("id")
        )
        if source_post_id is not None:
            dedupe_key = ("post_id", source_post_id)
        else:
            dedupe_key = ("content", content, timestamp, publish_time, content_type)
        if dedupe_key in seen_post_keys:
            continue
        seen_post_keys.add(dedupe_key)
        posts.append(
            UserPost(
                content=content,
                timestamp=timestamp,
                content_type=content_type,
                publish_time=publish_time,
                like_count=_safe_int(row.get("点赞数", 0)),
                comment_count=_safe_int(row.get("评论数", 0)),
                share_count=_safe_int(row.get("转发数", 0)),
                post_id=source_post_id,
            )
        )

    if not posts:
        raise ValueError(f"用户 {user_name} 的帖子数据为空，无法生成画像。")

    source = UserProfileSource(
        user_name=canonical_user_name,
        user_info=user_info,
        posts=posts,
    )
    return source, ""


def _build_json_schema_response_format(schema: Any) -> dict[str, Any]:
    """把 Pydantic schema 转成 OpenAI response_format=json_schema 结构。"""
    from langchain_core.utils.function_calling import convert_to_openai_function

    function = convert_to_openai_function(schema, strict=False)
    parameters = function.get("parameters") or {}
    return {
        "type": "json_schema",
        "json_schema": {
            "name": function.get("name") or getattr(schema, "__name__", "Output"),
            "schema": parameters,
            "strict": False,
        },
    }


def _build_llm_callable(llm: ChatOpenAI, use_json_schema: bool = True) -> LlmCallable:
    """把通用 LLM 客户端包装成画像生成器使用的调用签名。

    use_json_schema=True 时通过 response_format(json_schema) 引导模型输出固定结构，
    但解析统一走 json_repair（自动剥离 ```json 围栏、修复语法），
    不依赖 LangChain 的严格解析器；端点不支持 response_format 时回退纯文本解析。
    """
    from langchain_core.messages import HumanMessage, SystemMessage

    async def _call(
        system_prompt: str,
        user_prompt: str,
        schema: Any = None,
    ) -> dict[str, Any]:
        messages = [
            SystemMessage(content=system_prompt),
            HumanMessage(content=user_prompt),
        ]

        if use_json_schema and schema is not None:
            try:
                response_format = _build_json_schema_response_format(schema)
                structured_llm = llm.bind(response_format=response_format)
                result = await structured_llm.ainvoke(messages)
                extracted = _extract_json_dict(getattr(result, "content", None))
                if extracted:
                    return extracted
            except Exception as exc:  # noqa: BLE001
                print(f"[结构化输出] response_format 调用失败，回退纯文本解析：{exc}")

        result = await llm.ainvoke(messages)
        return _extract_json_dict(result.content)

    return _call


def _extract_json_dict(raw_content: Any) -> dict[str, Any]:
    """使用 json_repair 从模型原始输出中提取 JSON 对象。"""
    from json_repair import repair_json

    text = _normalize_text_content(raw_content)
    if not text:
        return {}
    try:
        payload = json.loads(repair_json(text))
        if isinstance(payload, dict):
            return payload
    except Exception:
        pass
    return {}


def _normalize_text_content(content: Any) -> str:
    """归一化模型返回文本。"""
    if isinstance(content, str):
        return content.strip()
    if isinstance(content, list):
        parts: list[str] = []
        for item in content:
            if isinstance(item, str):
                parts.append(item)
                continue
            if isinstance(item, dict):
                text = item.get("text") or item.get("content")
                if text:
                    parts.append(str(text))
                continue
            parts.append(str(item))
        return "\n".join(part.strip() for part in parts if part.strip())
    if content is None:
        return ""
    return str(content).strip()


def _load_json_object(path: Path) -> dict[str, Any]:
    """读取 JSON 对象。"""
    with path.open("r", encoding="utf-8") as file:
        payload = json.load(file)
    if not isinstance(payload, dict):
        raise ValueError(f"文件 {path} 必须是 JSON 对象。")
    return payload


def _load_tabular_rows(path: Path) -> list[dict[str, Any]]:
    """读取 Excel 或 CSV 表格数据。"""
    if not path.exists():
        raise FileNotFoundError(f"找不到数据文件：{path}")
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            reader = csv.DictReader(file)
            return [dict(row) for row in reader]
    if suffix != ".xlsx":
        raise ValueError(f"暂不支持的表格格式：{path.suffix}")

    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ModuleNotFoundError(
            "缺少 openpyxl，请先执行 `uv add openpyxl` 或 `uv sync`。"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    result: list[dict[str, Any]] = []
    for row in rows[1:]:
        item = {
            header[index]: row[index] if index < len(row) else None
            for index in range(len(header))
            if header[index]
        }
        if any(value not in (None, "") for value in item.values()):
            result.append(item)
    workbook.close()
    return result


def _validate_tabular_columns(
    rows: list[dict[str, Any]],
    path: Path,
    required_columns: tuple[str, ...],
    supported_columns: tuple[str, ...],
) -> None:
    """校验表格列是否符合旧版数据约定。"""
    if not rows:
        raise ValueError(f"数据文件为空：{path}")
    keys = {str(key).strip() for row in rows for key in row.keys() if str(key).strip()}
    missing_required = [column for column in required_columns if column not in keys]
    if missing_required:
        joined = "、".join(missing_required)
        raise ValueError(f"{path} 缺少必需列：{joined}")
    if not keys.intersection(supported_columns):
        joined = "、".join(supported_columns)
        raise ValueError(f"{path} 不符合预期列结构，应至少包含这些列中的一部分：{joined}")


def _extract_all_usernames(data_path: Path, user_file: str) -> list[str]:
    """从 user.xlsx/csv 中提取所有去重后的用户名列表。"""
    user_path = data_path / user_file
    rows = _load_tabular_rows(user_path)
    seen: set[str] = set()
    usernames: list[str] = []
    for row in rows:
        name = _normalize_username(row.get("用户名"))
        if name and name not in seen:
            seen.add(name)
            usernames.append(name)
    if not usernames:
        raise ValueError(f"在 {user_path} 中未找到任何用户名。")
    return usernames


def _normalize_username(value: Any) -> str:
    """统一用户名匹配规则。"""
    return _normalize_scalar(value).lower()


def _select_best_user_row(rows: list[dict[str, Any]]) -> dict[str, Any] | None:
    """同名用户存在多行时，优先选择信息最完整的一行。"""
    if not rows:
        return None
    scored_rows = sorted(
        rows,
        key=lambda row: (
            sum(1 for column in PORTRAIT_USER_SUPPORTED_COLUMNS if _normalize_scalar(row.get(column))),
            _safe_int(row.get("创建时间戳", 0)),
        ),
        reverse=True,
    )
    return scored_rows[0]


def _resolve_post_timestamp(row: dict[str, Any]) -> int:
    """优先使用时间戳，缺失时再解析可读时间。"""
    timestamp = _safe_int(row.get("发布时间戳", 0))
    if timestamp > 0:
        return timestamp
    publish_time = _normalize_scalar(row.get("发布时间"))
    if publish_time:
        return _parse_publish_time_to_timestamp(publish_time)
    return 0


def _collect_env_versions() -> dict[str, str]:
    """收集关键依赖版本（P2-E 可复现存档）。"""
    import numpy as np
    import pandas as pd
    import sklearn

    versions: dict[str, str] = {
        "python": os.sys.version.split()[0],
        "numpy": np.__version__,
        "pandas": pd.__version__,
        "scikit-learn": sklearn.__version__,
    }
    try:
        import optuna

        versions["optuna"] = optuna.__version__
    except ImportError:
        pass
    try:
        import sentence_transformers

        versions["sentence-transformers"] = sentence_transformers.__version__
    except ImportError:
        pass
    return versions




def _write_json(path: Path, payload: dict[str, Any]) -> None:
    """写入 JSON 文件（自动把 numpy 类型转成原生类型）。"""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        json.dump(payload, file, ensure_ascii=False, indent=2, default=_json_default)


def _json_default(value: Any) -> Any:
    """json.dump 的 default 序列化器：numpy 类型 → Python 原生类型。"""
    if isinstance(value, np.ndarray):
        return value.tolist()
    if isinstance(value, np.integer):
        return int(value)
    if isinstance(value, np.floating):
        return float(value)
    if isinstance(value, np.bool_):
        return bool(value)
    raise TypeError(f"Object of type {value.__class__.__name__} is not JSON serializable")


def _resolve_output_path(path_value: str | None, default_relative_path: str) -> Path:
    """解析输出路径。"""
    if path_value:
        output_path = Path(path_value)
    else:
        output_path = Path(default_relative_path)
    return output_path.resolve()


def _resolve_portrait_reference_time(reference_time: str | None) -> tuple[int, str]:
    """解析画像生成使用的参考时间。"""
    raw_value = reference_time.strip() if isinstance(reference_time, str) else ""
    while not raw_value:
        raw_value = input(
            "请输入画像生成参考时间（例如 2026-04-01 12:00:00，默认按 Asia/Shanghai 解析）："
        ).strip()
    return _parse_reference_time_to_timestamp(raw_value), raw_value


def _record_failed_portrait_user(
    output_path: Path,
    user_name: str,
    reference_time_text: str,
    error_message: str,
) -> Path:
    """记录画像生成失败的用户名单。"""
    report_path = output_path.parent / "failed_users.json"
    payload: dict[str, Any] = {
        "failed_users": [],
        "failed_details": [],
    }
    if report_path.exists():
        try:
            loaded = _load_json_object(report_path)
            if isinstance(loaded.get("failed_users"), list):
                payload["failed_users"] = list(loaded["failed_users"])
            if isinstance(loaded.get("failed_details"), list):
                payload["failed_details"] = list(loaded["failed_details"])
        except Exception:
            payload = {
                "failed_users": [],
                "failed_details": [],
            }

    if user_name not in payload["failed_users"]:
        payload["failed_users"].append(user_name)

    payload["failed_details"] = [
        item
        for item in payload["failed_details"]
        if not isinstance(item, dict) or item.get("user_name") != user_name
    ]
    payload["failed_details"].append(
        {
            "user_name": user_name,
            "reference_time": reference_time_text,
            "error_message": error_message,
        }
    )
    _write_json(report_path, payload)
    return report_path


def _safe_int(value: Any) -> int:
    """安全解析整数。"""
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0


def _safe_optional_int(value: Any) -> int | None:
    """安全解析可选整数。"""
    try:
        if value in (None, ""):
            return None
        return int(value)
    except (TypeError, ValueError):
        return None


def _safe_float(value: Any) -> float:
    """安全解析浮点数。"""
    try:
        if value in (None, ""):
            return 0.0
        text = str(value).strip()
        if text.lower() == "nan":
            return 0.0
        return float(text)
    except (TypeError, ValueError):
        return 0.0


def _normalize_scalar(value: Any) -> str:
    """归一化单元格文本。"""
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return text


def _parse_publish_time_to_timestamp(value: str) -> int:
    """尝试把可读时间解析成时间戳。"""
    text = value.strip()
    if not text:
        return 0
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ):
        try:
            return int(datetime.strptime(text, fmt).timestamp())
        except ValueError:
            continue
    return 0


def _parse_reference_time_to_timestamp(value: str) -> int:
    """解析画像生成参考时间。"""
    text = value.strip()
    if not text:
        raise ValueError("reference_time 不能为空。")
    if text.isdigit():
        timestamp = int(text)
        if timestamp > 0:
            return timestamp

    normalized = text.replace("Z", "+00:00")
    try:
        dt = datetime.fromisoformat(normalized)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=ZoneInfo("Asia/Shanghai"))
        return int(dt.timestamp())
    except ValueError:
        pass

    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d",
    ):
        try:
            dt = datetime.strptime(text, fmt).replace(tzinfo=ZoneInfo("Asia/Shanghai"))
            return int(dt.timestamp())
        except ValueError:
            continue
    raise ValueError(
        "reference_time 格式无效，请使用时间戳或类似 2026-04-01 12:00:00 的时间字符串。"
    )


def main() -> None:
    """脚本入口。"""
    try:
        from dotenv import load_dotenv

        _analysis_env = Path(__file__).resolve().parent / ".env"
        if _analysis_env.exists():
            load_dotenv(_analysis_env)
        else:
            load_dotenv()
    except ImportError:
        pass

    args = parse_args()
    if args.command == "portrait":
        if args.batch:
            asyncio.run(run_portrait_batch_command(args))
        elif not args.user_name:
            raise ValueError("请提供 --user-name，或使用 --batch 批量生成全部用户画像。")
        else:
            asyncio.run(run_portrait_command(args))
        return
    if args.command == "recommender":
        run_recommender_command(args)
        return
    if args.command == "retier":
        run_retier_command(args)
        return
    raise ValueError(f"不支持的命令：{args.command}")


if __name__ == "__main__":
    main()
