"""从 PHEME 转码后的 user.xlsx 生成 twitterapi.io 查询字符串。

对每个事件生成两份文件：
    handle_query.txt    "handle:cnnbrk OR handle:xxx OR ..."（按 handle 搜推文用）
    twitter_ids.txt     逗号分隔的用户 ID（批量查用户画像用，对应 data_pre.get_user_data）

用法：
    # 生成全部 9 个事件
    python scripts/generate_handle_queries.py

    # 只生成某个事件
    python scripts/generate_handle_queries.py --event ebola-essien

输入：data/pheme/processed/<event>/user.xlsx（列：用户名=用户ID，昵称=screen_name）
输出：data/pheme/processed/<event>/handle_query.txt 与 twitter_ids.txt
"""

from __future__ import annotations

import argparse
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    print("缺少 openpyxl，请先执行 `uv add openpyxl` 或 `uv sync`。")
    raise SystemExit(1)


def generate_for_event(event_dir: Path) -> tuple[int, str, str]:
    """读取 user.xlsx，返回 (用户数, handle_query, twitter_ids)。"""
    user_file = event_dir / "user.xlsx"
    if not user_file.is_file():
        raise SystemExit(f"找不到 {user_file}")

    wb = load_workbook(user_file, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]

    # 定位列
    def col_idx(name):
        for i, h in enumerate(header):
            if h == name:
                return i
        raise SystemExit(f"user.xlsx 缺少列：{name}")

    id_idx = col_idx("用户名")
    name_idx = col_idx("昵称")

    handles = []
    ids = []
    for r in rows[1:]:
        uid = str(r[id_idx]).strip() if r[id_idx] is not None else ""
        screen = str(r[name_idx]).strip() if r[name_idx] is not None else ""
        if screen and screen.lower() != "none":
            handles.append(screen)
        if uid:
            ids.append(uid)

    # 去重保序
    seen = set()
    handles_dedup = [h for h in handles if not (h in seen or seen.add(h))]
    seen_ids = set()
    ids_dedup = [i for i in ids if not (i in seen_ids or seen_ids.add(i))]

    handle_query = " OR ".join(f"handle:{h}" for h in handles_dedup)
    twitter_ids = ",".join(ids_dedup)

    return len(handles_dedup), handle_query, twitter_ids


def main() -> None:
    parser = argparse.ArgumentParser(description="生成 handle:xxx OR … 查询字符串")
    parser.add_argument("--event", default=None, help="指定事件名（默认全部）")
    parser.add_argument("--data-dir", default="data/pheme/processed", help="转码产物根目录")
    args = parser.parse_args()

    data_dir = Path(args.data_dir).resolve()
    if not data_dir.is_dir():
        raise SystemExit(f"目录不存在：{data_dir}")

    if args.event:
        events = [args.event]
    else:
        events = sorted(
            p.name for p in data_dir.iterdir() if (p / "user.xlsx").is_file()
        )

    for ev in events:
        ev_dir = data_dir / ev
        n, handle_query, twitter_ids = generate_for_event(ev_dir)

        (ev_dir / "handle_query.txt").write_text(handle_query, encoding="utf-8")
        (ev_dir / "twitter_ids.txt").write_text(twitter_ids, encoding="utf-8")

        print(f"[{ev}] {n} 个用户")
        print(f"    handle_query.txt : {handle_query[:120]}{'...' if len(handle_query) > 120 else ''}")
        print(f"    twitter_ids.txt  : {twitter_ids[:120]}{'...' if len(twitter_ids) > 120 else ''}")
        print()


if __name__ == "__main__":
    main()
