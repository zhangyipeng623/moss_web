"""从 国际传播 post.xlsx 生成 recommender 校准观测（无时间曲线版）。

post.xlsx 列：用户名,发文内容,帖子URL,发布时间,评论数,点赞数,转发数,图片,发文类型,发布时间戳
输出：data/国际传播/content_observations.json
  - story_id      ：从帖子 URL 提取的推文 ID
  - repost_count  ：转发数（校准目标）
  - view_count    ：转发数 + 点赞数 + 评论数（合成曝光量，作为缩放锚点）
  - text          ：发文内容

只保留“转发数 > 0”的帖子；view_count>100 等二次过滤交给 recommender 的 StoryManager。

用法：
    python scripts/build_content_observations.py \
        --post data/国际传播/post.xlsx \
        --out data/国际传播/content_observations.json
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    print("缺少 openpyxl，请先执行 `uv add openpyxl` 或 `uv sync`。")
    raise SystemExit(1)

_TWEET_ID_RE = re.compile(r"/status(?:es)?/(\d+)")


def _extract_tweet_id(url) -> str | None:
    m = _TWEET_ID_RE.search(str(url or ""))
    return m.group(1) if m else None


def _int(v) -> int:
    try:
        return int(v or 0)
    except (TypeError, ValueError):
        return 0


def main() -> None:
    parser = argparse.ArgumentParser(description="post.xlsx → 校准观测 JSON")
    parser.add_argument("--post", default="data/国际传播/post.xlsx", help="post.xlsx 路径")
    parser.add_argument("--out", default="data/国际传播/content_observations.json", help="输出 JSON")
    parser.add_argument("--min-repost", type=int, default=1, help="最小转发数")
    args = parser.parse_args()

    post_path = Path(args.post).resolve()
    if not post_path.is_file():
        raise SystemExit(f"找不到 {post_path}")

    wb = load_workbook(post_path, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    # 定位列
    idx = {name: i for i, name in enumerate(header)}
    need = ["发文内容", "帖子URL", "评论数", "点赞数", "转发数"]
    missing = [c for c in need if c not in idx]
    if missing:
        raise SystemExit(f"post.xlsx 缺少列：{missing}")

    records = []
    skipped_url = 0
    for row in rows:
        repost = _int(row[idx["转发数"]])
        if repost < args.min_repost:
            continue
        like = _int(row[idx["点赞数"]])
        comment = _int(row[idx["评论数"]])
        url = str(row[idx["帖子URL"]] or "")
        tid = _extract_tweet_id(url)
        if not tid:
            skipped_url += 1
            tid = f"idx_{len(records)}"  # URL 无推文ID时退化为序号
        records.append({
            "story_id": tid,
            "repost_count": repost,
            "view_count": repost + like + comment,
            "text": str(row[idx["发文内容"]] or ""),
        })

    out_path = Path(args.out).resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    payload = {"records": records, "anchor_percentile": 0.8, "max_iterations": 3}
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"生成 {len(records)} 条观测（转发数>={args.min_repost}）")
    print(f"  URL 无推文ID退化为序号的：{skipped_url}")
    print(f"写入：{out_path}")


if __name__ == "__main__":
    main()
